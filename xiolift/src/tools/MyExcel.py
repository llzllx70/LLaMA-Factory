
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from src.tools.MyDict import MyDict

class MyExcel:

    def __init__(self, file, log=None):
        self.sheet_ = None
        self.file = file
        self.wb = self.load()

        if log is not None:
            self.mylog = log
        else:
            self.mylog = print

    def load(self):
        return load_workbook(self.file, data_only=True) if os.path.exists(self.file) else None

    def record(self, ret):

        try:
            for row_num, llm_content in ret:
                self.sheet_[f'J{row_num}'] = llm_content['ret']

            self.wb.save(self.file)

        except Exception as e:
            self.mylog.error(f'{e.args[0]}')
            return []

    def load_excel_by_cols(self, sheet_name, cols, begin, end, contain_idx=True):
        """
        获取wb_[sheet_name] 的 cols 多行, 从第begin 开始, end 结束
        [[a, b], [c, d]]
        """

        try:
            self.sheet_ = self.wb[sheet_name]
            ret = []

            for index in range(begin, end+1):
                try:
                    ret_ = []
                    for col in cols:
                        k = self.sheet_.cell(index, col).value
                        if k is None:
                            k = 'None'
                        ret_.append(k)

                    if ret_ and len(ret_) == len(cols):

                        if contain_idx:
                            ret.append((index, ret_))  # 同时将行号也记录，用于结果记录

                        else:
                            ret.append(ret_)  # 同时将行号也记录，用于结果记录

                    index += 1

                except Exception as e:
                    self.mylog.info(f'{index} {e.args[0]}')

            return ret

        except Exception as e:
            self.mylog.error(f'{e.args[0]}')
            return []

        finally:
            self.wb.close()

    def load_excel_by_dict(self, dict_, contain_idx=False):

        return self.load_excel_by_cols(
            sheet_name=dict_['sheet'],
            cols=dict_['col'],
            begin=dict_['beg'],
            end=dict_['end'],
            contain_idx=contain_idx
        )

    def excel_2_stand_fanhua_json(self, dict_):

        """{'标准语义': [泛化语义]}"""

        ret = self.load_excel_by_dict(dict_)

        dict_ = MyDict()

        for a, b in ret:
            dict_.add_k_lv(b, a)

        return dict_

    def excel_2_fanhua_stand_list(self, dict_):

        """[[f1, f2...], [s1, s2....]]"""

        ret = self.load_excel_by_dict(dict_)

        ll = [[], []]

        for f, s in ret:
            ll[0].append(f)
            ll[1].append(s)

        return ll

    def excel_2_zh_en_ss(self, dict_):

        """[[f1, f2...], [e1, e2...], [s1, s2....]]"""

        ret = self.load_excel_by_dict(dict_)

        ll = [[], [], []]

        for z, e, s in ret:
            ll[0].append(z)
            ll[1].append(e)
            ll[2].append(s)

        return ll

    @staticmethod
    def open(file_, sheet_):

        if os.path.exists(file_):
            wb = load_workbook(file_, data_only=True)
        else:
            wb = Workbook()

        return wb

    @staticmethod
    def save(file_, sheet_, title_, width_, data_):

        print(f'save to: {file_} {sheet_}')
        wb = MyExcel.open(file_, sheet_)
        s = wb.create_sheet(sheet_, -1)

        c_ = 'A'
        for t, w in zip(title_, width_):

            s.column_dimensions[c_].width = w
            s[f'{c_}1'] = t
            c_ = chr(ord(c_) + 1)

        for i, row in enumerate(data_):
            for j, col in enumerate(row):
                c = chr(ord('A') + j)
                k = f'{c}{i + 2}'
                s[k] = str(col)

        wb.save(file_)
        wb.close()
